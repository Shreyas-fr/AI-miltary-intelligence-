import os
import streamlit as st
import pandas as pd
from io import BytesIO

# --- Authentication & Role Check ---
from utils.auth import require_auth
require_auth(['Analyst', 'Commander'])
# -----------------------------------

import importlib
import utils.doc_extraction
import utils.text_parsers
importlib.reload(utils.doc_extraction)
importlib.reload(utils.text_parsers)
from utils.doc_extraction import extract_intelligence_from_text, merge_extractions
from utils.text_parsers import parse_file

st.set_page_config(page_title="Document Intelligence", page_icon="📄", layout="wide")

def load_css(file_name: str):
    if os.path.exists(file_name):
        with open(file_name) as f:
            st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)

load_css("assets/style.css")

st.title("📄 | Document Intelligence Extraction")
st.markdown(
    "##### Automatically extract structured tactical data from uploaded intelligence reports. "
    "Features strict adherence to explicit facts only, rejecting hallucinated values."
)

st.divider()

st.info("⚠️ **Local-Only Feature:** This extraction pipeline runs entirely offline using Ollama and local text parsers. It will not execute in cloud deployments (e.g. Render).")

st.subheader("1. Upload Documents")
uploaded_files = st.file_uploader("Upload intelligence reports", type=["txt", "md", "pdf", "docx", "xlsx"], accept_multiple_files=True)

if "extracted_data" not in st.session_state:
    st.session_state["extracted_data"] = []

if st.button("Extract Intelligence", type="primary", disabled=not uploaded_files):
    all_results = []
    used_backends = set()
    error_occurred = False
    
    with st.spinner("Parsing documents with Ollama (Local)..."):
        for file in uploaded_files:
            file_bytes = file.read()
            filename = file.name
            try:
                # Parse text
                text = parse_file(filename, file_bytes)
                # Extract
                results, actual_backend = extract_intelligence_from_text(text, filename)
                all_results.extend(results)
                used_backends.add(actual_backend)
            except ConnectionError as ce:
                st.error(str(ce))
                st.stop()
            except Exception as e:
                st.error(f"Failed to process {filename}: {e}")
                error_occurred = True
                st.stop()
                
    if all_results:
        st.session_state["extracted_data"] = all_results
        st.session_state["backend_info"] = " & ".join(list(used_backends))
        st.success(f"Successfully extracted intelligence from {len(uploaded_files)} documents.")
    elif not error_occurred:
        st.warning("No intelligence entities found in the documents.")

st.divider()

st.subheader("3. Extracted Intelligence Database")

if st.session_state["extracted_data"]:
    st.info(f"Extracted via: **{st.session_state.get('backend_info', 'Unknown')}**")
    df_merged = merge_extractions(st.session_state["extracted_data"])
    
    st.dataframe(df_merged, use_container_width=True)
    
    # Generate Excel
    def to_excel(df: pd.DataFrame):
        output = BytesIO()
        
        # Use pandas with openpyxl engine
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Intelligence')
            
            # Formatting
            workbook = writer.book
            worksheet = writer.sheets['Intelligence']
            
            # Import styles
            from openpyxl.styles import Font
            grey_italic = Font(color="808080", italic=True)
            
            for row in worksheet.iter_rows(min_row=2, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
                for cell in row:
                    if cell.value == "Not found in document":
                        cell.font = grey_italic
                        
        processed_data = output.getvalue()
        return processed_data

    excel_data = to_excel(df_merged)
    
    st.download_button(
        label="📥 Export to Excel (.xlsx)",
        data=excel_data,
        file_name="Intelligence_Extraction.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary"
    )
else:
    st.info("Upload documents and extract to view the database.")
